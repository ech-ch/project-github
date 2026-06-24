"""
excel_to_bb.py
--------------
Liest BB_Standard_Import.xlsx und generiert building_blocks.json
nach dem SAGA.ch Building-Block-Schema.

Struktur jedes Sheets:
  Zeile 1-4: Beschreibung / Beispiel (wird übersprungen)
  Zeile 5:   Spaltenköpfe (ID*, PARENT_ID, ...)
  Zeile 6+:  Daten

Verwendung:
    python excel_to_bb.py
    python excel_to_bb.py --input BB_Standard_Import.xlsx
    python excel_to_bb.py --validate-only
    python excel_to_bb.py --output output/building_blocks.json
"""

import json
import sys
import argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Bitte ausführen: pip install openpyxl")
    sys.exit(1)


HEADER_ROW = 5
DATA_START  = 6

GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
BLUE   = "\033[94m"
RESET  = "\033[0m"
BOLD   = "\033[1m"

def ok(msg):   print(f"  {GREEN}✓{RESET} {msg}")
def err(msg):  print(f"  {RED}✗{RESET} {msg}")
def warn(msg): print(f"  {YELLOW}⚠{RESET} {msg}")
def info(msg): print(f"  {BLUE}→{RESET} {msg}")


def clean(value) -> str:
    if value is None:
        return ""
    s = str(value).strip().replace("\n", " ").replace("\r", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def read_sheet(ws) -> list[dict]:
    headers = []
    for cell in ws[HEADER_ROW]:
        h = clean(cell.value).upper().replace("*", "").strip()
        headers.append(h)

    rows = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = clean(val)
        rows.append(row_dict)
    return rows


def build_scope(s1: str, s2: str, s3: str) -> list[str]:
    scope = []
    if s1.upper() == "Y": scope.append("S1")
    if s2.upper() == "Y": scope.append("S2")
    if s3.upper() == "Y": scope.append("S3")
    return scope


def validate(makros, mesos, mikros, variants, alternatives, references, assessments):
    errors = []
    makro_ids   = {m["ID"] for m in makros if m.get("ID")}
    meso_ids    = {m["ID"] for m in mesos if m.get("ID")}
    mikro_ids   = {m["ID"] for m in mikros if m.get("ID")}
    variant_ids = {v["VARIANT_ID"] for v in variants if v.get("VARIANT_ID")}
    alt_ids     = {a["ALTERNATIVE_ID"] for a in alternatives if a.get("ALTERNATIVE_ID")}

    for m in makros:
        for f in ["ID", "LEVEL", "TITLE", "SEMANTIC_SUMMARY"]:
            if not m.get(f):
                errors.append(f"MAKRO '{m.get('ID','?')}': Pflichtfeld '{f}' fehlt")

    for m in mesos:
        for f in ["ID", "PARENT_ID", "LEVEL", "TITLE", "SEMANTIC_SUMMARY"]:
            if not m.get(f):
                errors.append(f"MESO '{m.get('ID','?')}': Pflichtfeld '{f}' fehlt")
        if m.get("PARENT_ID") and m["PARENT_ID"] not in makro_ids:
            errors.append(f"MESO '{m.get('ID')}': PARENT_ID '{m['PARENT_ID']}' ist keine gültige Makro-ID")

    for m in mikros:
        for f in ["ID", "PARENT_ID", "LEVEL", "TITLE", "SEMANTIC_SUMMARY"]:
            if not m.get(f):
                errors.append(f"MIKRO '{m.get('ID','?')}': Pflichtfeld '{f}' fehlt")
        if m.get("PARENT_ID") and m["PARENT_ID"] not in meso_ids:
            errors.append(f"MIKRO '{m.get('ID')}': PARENT_ID '{m['PARENT_ID']}' ist keine gültige Meso-ID")

    for v in variants:
        if not v.get("VARIANT_ID"):
            errors.append("MIKRO_VARIANTS: Zeile ohne VARIANT_ID")
        elif v.get("MICRO_ID") and v["MICRO_ID"] not in mikro_ids:
            errors.append(f"VARIANT '{v['VARIANT_ID']}': MICRO_ID '{v['MICRO_ID']}' nicht gefunden")

    for a in alternatives:
        if not a.get("ALTERNATIVE_ID"):
            errors.append("MIKRO_ALTERNATIVES: Zeile ohne ALTERNATIVE_ID")
        elif a.get("MICRO_ID") and a["MICRO_ID"] not in mikro_ids:
            errors.append(f"ALTERNATIVE '{a['ALTERNATIVE_ID']}': MICRO_ID '{a['MICRO_ID']}' nicht gefunden")

    for i, r in enumerate(references, 1):
        targets = [bool(r.get("TARGET_MICRO_ID")), bool(r.get("TARGET_VARIANT_ID")), bool(r.get("TARGET_ALTERNATIVE_ID"))]
        if sum(targets) == 0:
            errors.append(f"MIKRO_REFERENCES Zeile {i+DATA_START-1}: Keine Target-Spalte gesetzt")
        if sum(targets) > 1:
            errors.append(f"MIKRO_REFERENCES Zeile {i+DATA_START-1}: Mehrere Target-Spalten gesetzt (nur eine erlaubt)")
        if r.get("TARGET_MICRO_ID") and r["TARGET_MICRO_ID"] not in mikro_ids:
            errors.append(f"MIKRO_REFERENCES Zeile {i+DATA_START-1}: TARGET_MICRO_ID '{r['TARGET_MICRO_ID']}' nicht gefunden")
        if r.get("TARGET_VARIANT_ID") and r["TARGET_VARIANT_ID"] not in variant_ids:
            errors.append(f"MIKRO_REFERENCES Zeile {i+DATA_START-1}: TARGET_VARIANT_ID '{r['TARGET_VARIANT_ID']}' nicht gefunden")
        if r.get("TARGET_ALTERNATIVE_ID") and r["TARGET_ALTERNATIVE_ID"] not in alt_ids:
            errors.append(f"MIKRO_REFERENCES Zeile {i+DATA_START-1}: TARGET_ALTERNATIVE_ID '{r['TARGET_ALTERNATIVE_ID']}' nicht gefunden")

    for i, a in enumerate(assessments, 1):
        targets = [bool(a.get("TARGET_MICRO_ID")), bool(a.get("TARGET_VARIANT_ID")), bool(a.get("TARGET_ALTERNATIVE_ID"))]
        if sum(targets) == 0:
            errors.append(f"MIKRO_ASSESSMENTS Zeile {i+DATA_START-1}: Keine Target-Spalte gesetzt")
        if sum(targets) > 1:
            errors.append(f"MIKRO_ASSESSMENTS Zeile {i+DATA_START-1}: Mehrere Target-Spalten gesetzt (nur eine erlaubt)")

    return errors


def build_reference(ref: dict) -> dict:
    r = {"label": ref.get("LABEL", "")}
    if ref.get("ORGANIZATION"):
        r["organization"] = ref["ORGANIZATION"]
    if ref.get("REFERENCE_TYPE"):
        r["reference_type"] = ref["REFERENCE_TYPE"].lower()
    if ref.get("URL"):
        r["url"] = ref["URL"]
    return r


def build_assessment(ass: dict) -> dict:
    return {
        "interface_scope":  build_scope(
            ass.get("SCOPE_S1", "N"),
            ass.get("SCOPE_S2", "N"),
            ass.get("SCOPE_S3", "N"),
        ),
        "normative_status": ass.get("NORMATIVE_STATUS", "").lower(),
    }


def assemble(makros, mesos, mikros, variants, alternatives, references, assessments) -> dict:
    # Lookup-Tabellen
    refs_by_micro   = defaultdict(list)
    refs_by_variant = defaultdict(list)
    refs_by_alt     = defaultdict(list)
    for r in references:
        ref_obj = build_reference(r)
        if r.get("TARGET_MICRO_ID"):
            refs_by_micro[r["TARGET_MICRO_ID"]].append(ref_obj)
        elif r.get("TARGET_VARIANT_ID"):
            refs_by_variant[r["TARGET_VARIANT_ID"]].append(ref_obj)
        elif r.get("TARGET_ALTERNATIVE_ID"):
            refs_by_alt[r["TARGET_ALTERNATIVE_ID"]].append(ref_obj)

    ass_by_micro   = defaultdict(list)
    ass_by_variant = defaultdict(list)
    ass_by_alt     = defaultdict(list)
    for a in assessments:
        ass_obj = build_assessment(a)
        if a.get("TARGET_MICRO_ID"):
            ass_by_micro[a["TARGET_MICRO_ID"]].append(ass_obj)
        elif a.get("TARGET_VARIANT_ID"):
            ass_by_variant[a["TARGET_VARIANT_ID"]].append(ass_obj)
        elif a.get("TARGET_ALTERNATIVE_ID"):
            ass_by_alt[a["TARGET_ALTERNATIVE_ID"]].append(ass_obj)

    vars_by_micro = defaultdict(list)
    for v in variants:
        vid = v.get("VARIANT_ID", "")
        entry = {"variant_id": vid, "title": v.get("TITLE", "")}
        if refs_by_variant.get(vid):
            entry["references"] = refs_by_variant[vid]
        if ass_by_variant.get(vid):
            entry["assessments"] = ass_by_variant[vid]
        vars_by_micro[v.get("MICRO_ID", "")].append(entry)

    alts_by_micro = defaultdict(list)
    for a in alternatives:
        aid = a.get("ALTERNATIVE_ID", "")
        entry = {"alternative_id": aid, "title": a.get("TITLE", "")}
        if refs_by_alt.get(aid):
            entry["references"] = refs_by_alt[aid]
        if ass_by_alt.get(aid):
            entry["assessments"] = ass_by_alt[aid]
        alts_by_micro[a.get("MICRO_ID", "")].append(entry)

    mikros_by_meso = defaultdict(list)
    for mi in mikros:
        mid = mi["ID"]
        block = {
            "id":               mid,
            "parent_id":        mi.get("PARENT_ID") or None,
            "level":            mi.get("LEVEL", "MIKRO"),
            "title":            mi.get("TITLE", ""),
            "semantic_summary": mi.get("SEMANTIC_SUMMARY", ""),
            "variants":         vars_by_micro.get(mid, []),
            "alternatives":     alts_by_micro.get(mid, []),
            "references":       refs_by_micro.get(mid, []),
            "assessments":      ass_by_micro.get(mid, []),
        }
        mikros_by_meso[mi.get("PARENT_ID", "")].append(block)

    mesos_by_makro = defaultdict(list)
    for me in mesos:
        meid = me["ID"]
        block = {
            "id":               meid,
            "parent_id":        me.get("PARENT_ID") or None,
            "level":            me.get("LEVEL", "MESO"),
            "title":            me.get("TITLE", ""),
            "semantic_summary": me.get("SEMANTIC_SUMMARY", ""),
            "mikro_blocks":     mikros_by_meso.get(meid, []),
        }
        mesos_by_makro[me.get("PARENT_ID", "")].append(block)

    makro_blocks = []
    for ma in makros:
        maid = ma["ID"]
        makro_blocks.append({
            "id":               maid,
            "parent_id":        ma.get("PARENT_ID") or None,
            "level":            ma.get("LEVEL", "MAKRO"),
            "title":            ma.get("TITLE", ""),
            "semantic_summary": ma.get("SEMANTIC_SUMMARY", ""),
            "meso_blocks":      mesos_by_makro.get(maid, []),
        })

    return {
        "document": {
            "standard_id":    "eCH-0014",
            "standard_title": "SAGA.ch",
            "version":        "8.0",
        },
        "building_blocks": makro_blocks,
    }


def main():
    parser = argparse.ArgumentParser(description="BB_Standard_Import.xlsx → building_blocks.json")
    parser.add_argument("--input",         type=Path, default=Path("BB_Standard_Import.xlsx"))
    parser.add_argument("--output",        type=Path, default=Path("building_blocks.json"))
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()

    print(f"\n{BOLD}{'='*55}{RESET}")
    print(f"{BOLD}  SAGA.ch Building Blocks – Excel → JSON{RESET}")
    print(f"{BOLD}{'='*55}{RESET}\n")

    if not args.input.exists():
        print(f"{RED}Fehler: '{args.input}' nicht gefunden.{RESET}")
        sys.exit(1)

    info(f"Lese: {args.input}")

    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(args.input)
    except Exception as e:
        print(f"{RED}Fehler beim Öffnen: {e}{RESET}")
        sys.exit(1)

    makros       = read_sheet(wb["MAKRO"])
    mesos        = read_sheet(wb["MESO"])
    mikros       = read_sheet(wb["MIKRO"])
    variants     = read_sheet(wb["MIKRO_VARIANTS"])
    alternatives = read_sheet(wb["MIKRO_ALTERNATIVES"])
    references   = read_sheet(wb["MIKRO_REFERENCES"])
    assessments  = read_sheet(wb["MIKRO_ASSESSMENTS"])

    print(f"\n{BOLD}Gelesene Zeilen:{RESET}")
    print(f"  MAKRO:              {len(makros)}")
    print(f"  MESO:               {len(mesos)}")
    print(f"  MIKRO:              {len(mikros)}")
    print(f"  MIKRO_VARIANTS:     {len(variants)}")
    print(f"  MIKRO_ALTERNATIVES: {len(alternatives)}")
    print(f"  MIKRO_REFERENCES:   {len(references)}")
    print(f"  MIKRO_ASSESSMENTS:  {len(assessments)}")

    print(f"\n{BOLD}Validierung:{RESET}")
    errors = validate(makros, mesos, mikros, variants, alternatives, references, assessments)

    if errors:
        print(f"\n{RED}{BOLD}  {len(errors)} Fehler:{RESET}")
        for e in errors:
            err(e)
        if args.validate_only:
            sys.exit(1)
    else:
        ok("Keine Fehler – alle Referenzen valide.")

    if args.validate_only:
        print(f"\n{YELLOW}--validate-only: kein Output.{RESET}\n")
        return

    result = assemble(makros, mesos, mikros, variants, alternatives, references, assessments)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    makro_count = len(result["building_blocks"])
    meso_count  = sum(len(ma["meso_blocks"]) for ma in result["building_blocks"])
    mikro_count = sum(len(me["mikro_blocks"]) for ma in result["building_blocks"] for me in ma["meso_blocks"])

    print(f"\n{BOLD}Output:{RESET}")
    ok(f"{args.output.resolve()}")
    print(f"\n{BOLD}JSON-Struktur:{RESET}")
    print(f"  Makro-Blöcke:  {makro_count}")
    print(f"  Meso-Blöcke:   {meso_count}")
    print(f"  Mikro-Blöcke:  {mikro_count}\n")


if __name__ == "__main__":
    main()
